"""
Fill the GHGA submission Excel workbook with available metadata.

NOTE: The GHGA submission this script prepared is complete, and its inputs
(participant ethnicity, sex, age, birth country) contained identifying
metadata that has been removed from this repository -- see
doc/data_submission/ (removed) and the commit that dropped it. This script
is kept only for provenance/reference and is not runnable as-is; it is not
called from any Snakemake rule.

Reads (no longer present in this repo):
  data/datasets.yml                                  (authoritative run list)
  doc/data_submission/ethnic_background_fixed.tsv    (ancestry, sex, country)
  doc/tables/samples_metadata.tsv                    (age, birth_country)

Writes all available metadata into:
  doc/data_submission/ghga_submission_full.xlsx

Publication scope: T2T00–T2T19 (excl T2T14), 23 individuals / 25 sub-samples.
Excluded: run #06048 (aborted T2T00 PoreC run).

Fields left blank require manual completion:
  - DataAccessCommittee email       (pending GHGA helpdesk)
  - DataAccessPolicy policy_text    (pending legal review)
  - Assembly / VCF file names       (to be verified once files are finalised)
  - workflow_doi / publication doi  (after publication)
"""

import re
import yaml
from pathlib import Path

import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE          = Path(__file__).resolve().parents[2]
DATASETS_FILE = BASE / "data/datasets.yml"
ETHNIC_FILE   = BASE / "doc/data_submission/ethnic_background_fixed.tsv"
META_FILE     = BASE / "doc/tables/samples_metadata.tsv"
TEMPLATE_FILE = BASE / "doc/data_submission/ghga_submission_full.xlsx"
OUTPUT_FILE   = BASE / "data/upload_folder/GHGA/ghga_submission_ONT_T2T_v3.xlsx"

HEADER_ROWS  = 6   # rows 1-6 are template headers in each GHGA sheet

# Publication sample list (T2T00–T2T19, excl T2T14, with family sub-samples)
PUBLICATION_SAMPLES = [
    "T2T00", "T2T00_1", "T2T00_2",
    "T2T01", "T2T02", "T2T03",
    "T2T04", "T2T04_1", "T2T04_2",
    "T2T05", "T2T06", "T2T07", "T2T08", "T2T09", "T2T10",
    "T2T11", "T2T12", "T2T13",
    "T2T15", "T2T16", "T2T17", "T2T18", "T2T19",
]

EXCLUDE_RUNS = {"06048"}   # aborted T2T00 PoreC run

DATASET_ALIAS         = "dataset_T2T_ONT"
STUDY_ALIAS           = "study_T2T_ONT"
DAP_ALIAS             = "dap_T2T_ONT"
DAC_ALIAS             = "dac_IMGAG"
ANALYSIS_METHOD_ALIAS = "method_T2T_ONT_workflow"

# ---------------------------------------------------------------------------
# Controlled-vocabulary lookup tables
# ---------------------------------------------------------------------------

ETHNICITY_TO_HANCESTRO = {
    "German":                               ("European",        "HANCESTRO:0005"),
    "German (Baden-Württemberg/Latvia)":    ("European",        "HANCESTRO:0005"),
    "German (Bavarian/Sudeten/Romanian)":   ("European",        "HANCESTRO:0005"),
    "German (Rhineland-Palatinate/Berlin)": ("European",        "HANCESTRO:0005"),
    "German (Saarland)":                    ("European",        "HANCESTRO:0005"),
    "German (Mixed regions)":               ("European",        "HANCESTRO:0005"),
    "German (Saxony-Anhalt)":               ("European",        "HANCESTRO:0005"),
    "German/Austrian":                      ("European",        "HANCESTRO:0005"),
    "German/Polish (Saxony-Anhalt)":        ("European",        "HANCESTRO:0005"),
    "Austrian":                             ("European",        "HANCESTRO:0005"),
    "French":                               ("European",        "HANCESTRO:0005"),
    "Spanish":                              ("European",        "HANCESTRO:0005"),
    "Croatian":                             ("European",        "HANCESTRO:0005"),
    "Russian":                              ("European",        "HANCESTRO:0005"),
    "Eastern Slavic/German":                ("European",        "HANCESTRO:0005"),
    "Eastern European Mix":                 ("European",        "HANCESTRO:0005"),
    "South Indian":                         ("South Asian",     "HANCESTRO:0006"),
    "Taiwanese (Han Chinese)":              ("East Asian",      "HANCESTRO:0009"),
    "Vietnamese":                           ("Southeast Asian", "HANCESTRO:0028"),
    "Chilean":                              ("Latin American",  "HANCESTRO:0014"),
    "Mexican":                              ("Latin American",  "HANCESTRO:0014"),
}

COUNTRY_MAP = {
    "Germany": ("Germany", "NCIT:C16636"),
    "Austria": ("Austria", "NCIT:C16312"),
    "Spain":   ("Spain",   "NCIT:C17152"),
    "Russia":  ("Russia",  "NCIT:C17147"),
    "France":  ("France",  "NCIT:C16607"),
    "Chile":   ("Chile",   "NCIT:C16453"),
    "India":   ("India",   "NCIT:C16683"),
    "Mexico":  ("Mexico",  "NCIT:C16888"),
    "Poland":  ("Poland",  "NCIT:C17062"),
    "Taiwan":  ("Taiwan",  "NCIT:C17222"),
    "Croatia": ("Croatia", "NCIT:C16490"),
    "Vietnam": ("Vietnam", "NCIT:C17272"),
}

SEX_MAP = {
    "male":   "MALE",
    "female": "FEMALE",
    "n/a":    "UNKNOWN",
    "":       "UNKNOWN",
}

# datasets.yml key → ExperimentMethod alias
# HQ_paternal / HQ_maternal: owned by sub-samples (T2T00_1, T2T04_1 etc.), skipped here
# HQ_herro: only the extra parent-sample entries (not in UL list) are LSK
# HQ_duplex: same physical flowcells, not separately submitted
KEY_TO_METHOD = {
    "UL":    "method_ULK114",
    "POREC": "method_PoreC",
    "APK":   "method_APK114",
}

KEY_TO_TYPE = {
    "UL":       "UL",
    "POREC":    "POREC",
    "APK":      "APK",
    "HQ_HERRO": "LSK",
}

# ExperimentMethod full definitions
METHOD_DEFS = {
    "method_ULK114": {
        "name":        "UL-LR-ONT-SQK-ULK114",
        "library_type":"WGS",
        "kit_name":    "OTHER",   # SQK-ULK114 not in GHGA CV; actual kit: Ultra-Long DNA Sequencing Kit V14
        "description": "Oxford Nanopore ultra-long whole-genome sequencing using SQK-ULK114",
    },
    "method_LSK114": {
        "name":        "LR-ONT-SQK-LSK114",
        "library_type":"WGS",
        "kit_name":    "OTHER",   # SQK-LSK114 not in GHGA CV; actual kit: Ligation Sequencing Kit V14
        "description": "Oxford Nanopore whole-genome sequencing using SQK-LSK114",
    },
    "method_APK114": {
        "name":        "APK-ONT-SQK-APK114",
        "library_type":"WGS",
        "kit_name":    "OTHER",   # SQK-APK114 not in GHGA CV; actual kit: Assembly Polishing Kit V14
        "description": "Oxford Nanopore assembly-polishing sequencing using SQK-APK114",
    },
    "method_PoreC": {
        "name":        "PoreC-ONT-SQK-LSK114",
        "library_type":"CHROMOSOME_CONFORMATION_CAPTURE",
        "kit_name":    "OTHER",   # SQK-LSK114 not in GHGA CV; actual kit: ONT Pore-C Sequencing Kit
        "description": "Oxford Nanopore Pore-C chromatin conformation capture using SQK-LSK114",
    },
}


def age_to_5year(age) -> str:
    """Map age to GHGA 5-year bin enum values (transpiler uppercases these)."""
    if pd.isna(age):
        return "unknown"
    age = int(age)
    if age <=  5:  return "0_to_5"
    if age <= 10:  return "6_to_10"
    if age <= 15:  return "11_to_15"
    if age <= 20:  return "16_to_20"
    if age <= 25:  return "21_to_25"
    if age <= 30:  return "26_to_30"
    if age <= 35:  return "31_to_35"
    if age <= 40:  return "36_to_40"
    if age <= 45:  return "41_to_45"
    if age <= 50:  return "46_to_50"
    if age <= 55:  return "51_to_55"
    if age <= 60:  return "56_to_60"
    if age <= 65:  return "61_to_65"
    if age <= 70:  return "66_to_70"
    if age <= 75:  return "71_to_75"
    if age <= 80:  return "76_to_80"
    return "81_or_older"


def run_id_from_path(path: str) -> str:
    """Extract 5-digit run ID from a datasets.yml path entry."""
    return path.rstrip("/").split("/")[-1].split("_")[-1]


# ---------------------------------------------------------------------------
# Load source data
# ---------------------------------------------------------------------------
with open(DATASETS_FILE) as f:
    datasets = yaml.safe_load(f)

ethnic = pd.read_csv(ETHNIC_FILE, sep="\t")
ethnic.columns = [c.strip() for c in ethnic.columns]
ethnic = ethnic[ethnic["T2T Number"].isin(PUBLICATION_SAMPLES)].set_index("T2T Number")

meta = pd.read_csv(META_FILE, sep="\t")
meta = meta[meta["t2t_identifier"].isin(PUBLICATION_SAMPLES)].set_index("t2t_identifier")

# ---------------------------------------------------------------------------
# Build BAM filename lookup: run_id → filename on disk
# ---------------------------------------------------------------------------
basecalled_root = BASE / "data/basecalled"
fc_to_filename: dict[str, str] = {}
fc_to_path: dict[str, Path] = {}
for p in basecalled_root.glob("**/*.unmapped.bam"):
    # Filenames: 25006LRa010_04919.apk.unmapped.bam  or  …_05375.sup.unmapped.bam
    stem_parts = p.name.split(".")
    run_part = stem_parts[0]          # e.g. "24070LRa018_T2T00_PoreC_05375"
    fc = run_part.split("_")[-1]      # e.g. "05375"
    fc_to_filename[fc] = p.name
    fc_to_path[fc] = p

# ---------------------------------------------------------------------------
# Build entity tables
# ---------------------------------------------------------------------------

# --- Study ------------------------------------------------------------------
study_rows = [{
    "alias":        STUDY_ALIAS,
    "title":        (
        "Single-Platform Nanopore Sequencing Enables Diploid "
        "Telomere-to-Telomere Genome Assembly and Haplotype-Resolved 3D Chromatin Maps"
    ),
    "description":  (
        "Whole-genome sequencing of 23 individuals using Oxford Nanopore Technology "
        "ultra-long reads, Pore-C, and APK reads for haplotype-resolved near-telomere-to-telomere "
        "assembly and epigenomic analysis."
    ),
    "types":        "Whole Genome Sequencing",
    "affiliations": (
        "Institute of Medical Genetics and Applied Genomics, University of Tübingen; "
        "Institute for Bioinformatics and Medical Informatics (IBMI), University of Tübingen; "
        "NGS Competence Center Tübingen (NCCT), University of Tübingen"
    ),
}]

# --- Individuals & Samples --------------------------------------------------
individual_rows = []
sample_rows = []

for t2t_id in PUBLICATION_SAMPLES:
    if t2t_id not in ethnic.index:
        print(f"  WARNING: {t2t_id} not in ethnic_background_fixed.tsv, skipping individual")
        continue

    eth = ethnic.loc[t2t_id]
    sex = SEX_MAP.get(str(eth.get("Gender", "")).strip().lower(), "UNKNOWN")

    # Birth country: samples_metadata first, ethnic_background as fallback
    country_raw = ""
    if t2t_id in meta.index:
        country_raw = str(meta.loc[t2t_id, "birth_country"]).strip()
    if not country_raw or country_raw == "nan":
        country_raw = str(eth.get("Country", "")).strip()
    country_term, country_id = COUNTRY_MAP.get(country_raw, ("", ""))

    ethnicity_raw = str(eth.get("Ethnicity", "")).strip()
    anc_term, anc_id = ETHNICITY_TO_HANCESTRO.get(ethnicity_raw, ("", ""))

    age_decade = ""
    if t2t_id in meta.index:
        age_decade = age_to_5year(meta.loc[t2t_id, "age"])

    individual_rows.append({
        "alias":                     t2t_id,
        "sex":                       sex,
        "geographical_region_term":  country_term,
        "geographical_region_id":    country_id,
        "ancestry_terms":            anc_term,
        "ancestry_ids":              anc_id,
        "phenotypic_features_terms": "Unaffected",
        "phenotypic_features_ids":   "HP:0003674",
    })

    sample_rows.append({
        "alias":                                f"sample_{t2t_id}",
        "individual":                           t2t_id,
        "name":                                 f"{t2t_id}_blood_DNA",
        "description":                          (
            "Peripheral blood-derived high-molecular-weight DNA from a healthy adult"
        ),
        "case_control_status":                  "CONTROL",
        "disease_or_healthy":                   "HEALTHY",
        "biospecimen_type":                     "WHOLE_BLOOD",
        "biospecimen_tissue_term":              "blood",
        "biospecimen_tissue_id":                "BTO:0000089",
        "biospecimen_vital_status_at_sampling": "ALIVE",
        "biospecimen_age_at_sampling":          age_decade,
    })

# --- ExperimentMethods ------------------------------------------------------
method_rows = []
for alias, defn in METHOD_DEFS.items():
    method_rows.append({
        "alias":                              alias,
        "name":                               defn["name"],
        "description":                        defn["description"],
        "type":                               "Sequencing",
        "library_type":                       defn["library_type"],
        "library_selection_methods":          "RANDOM_METHOD",
        "library_preparation":                "Ligation-based",
        "library_preparation_kit_retail_name": defn["kit_name"],
        "library_preparation_kit_manufacturer":"Oxford Nanopore Technologies",
        "instrument_model":                   "PromethION",
        "sequencing_layout":                  "SE",
        "sequencing_center":                  (
            "Institute of Medical Genetics and Applied Genomics, Tübingen"
        ),
        "flow_cell_type": "PromethION",
    })

# --- Experiments & ResearchDataFiles ----------------------------------------
# Each physical flowcell run is one Experiment + one ResearchDataFile.
# Source of truth: datasets.yml UL / POREC / APK keys + LSK extras in HQ_herro.
# HQ_paternal / HQ_maternal / HQ_duplex are skipped (same physical files as
# sub-samples' own entries, or not submitted).

experiment_rows = []
file_rows = []
seen_run_ids: set[str] = set()        # global dedup across samples

# Track which run_ids belong to each sample (for analysis inputs later)
sample_own_runs: dict[str, set[str]] = {s: set() for s in PUBLICATION_SAMPLES}
# Also track ALL run_ids referenced by each sample (including cross-sample HQ_paternal etc.)
sample_all_runs: dict[str, set[str]] = {s: set() for s in PUBLICATION_SAMPLES}
# For file list generation: run_id → (sample, bam_type)
rid_to_sample_type: dict[str, tuple[str, str]] = {}

for t2t_id in PUBLICATION_SAMPLES:
    sample_data = datasets.get(t2t_id, {})
    if not isinstance(sample_data, dict):
        continue

    # UL run IDs for this sample (needed to detect extra LSK in HQ_herro)
    ul_ids: set[str] = set()
    if "UL" in sample_data:
        paths = sample_data["UL"] if isinstance(sample_data["UL"], list) else [sample_data["UL"]]
        ul_ids = {run_id_from_path(p) for p in paths}

    for key, paths in sample_data.items():
        if isinstance(paths, str):
            paths = [paths]
        key_up = key.upper()

        # Determine method alias for this key
        if key_up in KEY_TO_METHOD:
            method_alias = KEY_TO_METHOD[key_up]
            run_paths = paths
        elif key_up == "HQ_HERRO":
            # Only the extra entries not in UL (these are parent-sample LSK runs)
            run_paths = [p for p in paths if run_id_from_path(p) not in ul_ids]
            method_alias = "method_LSK114"
        else:
            # HQ_paternal, HQ_maternal, HQ_duplex — collect for analysis inputs but
            # don't create new experiments (already owned by sub-sample entries)
            for p in paths:
                rid = run_id_from_path(p)
                if rid not in EXCLUDE_RUNS:
                    sample_all_runs[t2t_id].add(rid)
            continue

        for path in run_paths:
            rid = run_id_from_path(path)
            if rid in EXCLUDE_RUNS:
                continue

            # Track for analysis inputs
            sample_own_runs[t2t_id].add(rid)
            sample_all_runs[t2t_id].add(rid)

            # Only create experiment + file once (global dedup for parent-sample LSK)
            if rid in seen_run_ids:
                continue
            seen_run_ids.add(rid)
            rid_to_sample_type[rid] = (t2t_id, KEY_TO_TYPE.get(key_up, ""))

            filename = fc_to_filename.get(rid, f"MISSING_{rid}.unmapped.bam")

            experiment_rows.append({
                "alias":             f"exp_{rid}",
                "title":             f"{t2t_id} {method_alias}",
                "experiment_method": method_alias,
                "sample":            f"sample_{t2t_id}",
                "type":              "Sequencing",
            })

            file_rows.append({
                "alias":                  f"file_{rid}",
                "name":                   filename,
                "format":                 "UBAM",
                "technical_replicate":    1,
                "experiments":            f"exp_{rid}",
                "dataset":                DATASET_ALIAS,
                "included_in_submission": True,
            })

# --- AnalysisMethod ---------------------------------------------------------
analysis_method_rows = [{
    "alias":              ANALYSIS_METHOD_ALIAS,
    "name":               "T2T-ONT Snakemake Workflow",
    "description":        (
        "Haplotype-resolved near-T2T assembly from ONT reads using Verkko graph assembly, "
        "HERRO error correction, Pore-C scaffolding and phasing, and Medaka APK polishing. "
        "Variant calling against T2T-CHM13v2.0 using dipcall and hapdiff."
    ),
    "type":               "Genome Assembly",
    "workflow_name":      "T2T-ONT",
    "workflow_version":   "",
    "workflow_repository":"https://github.com/imgag/T2T_ONT",
    "workflow_doi":       "https://doi.org/10.64898/2026.03.19.712851",
    "reference_name":     "T2T-CHM13v2.0",
    "reference_type":     "GENOME",
    "reference_source":   "NCBI",
    "reference_link":     "https://www.ncbi.nlm.nih.gov/datasets/genome/GCF_009914755.1/",
    "reference_version":  "2.0",
    "software_versions":  (
        "Verkko=2.2; HERRO=1.0; Hifiasm=0.25.0; Medaka=2.0; "
        "WhatsHap=2.3; Modkit=0.5; dipcall=0.3; hapdiff=1.2; Dorado=1.4"
    ),
}]

# --- Analyses (one per individual) ------------------------------------------
analysis_rows = []
for t2t_id in PUBLICATION_SAMPLES:
    # All run_ids referenced in this sample's assembly (own + cross-sample inputs)
    input_file_aliases = sorted(
        f"file_{rid}" for rid in sample_all_runs[t2t_id]
        if rid in seen_run_ids   # only include files that were actually registered
    )
    analysis_rows.append({
        "alias":               f"analysis_{t2t_id}",
        "analysis_method":     ANALYSIS_METHOD_ALIAS,
        "title":               f"T2T assembly and variant calling: {t2t_id}",
        "description":         (
            f"Haplotype-resolved T2T assembly and phased variant calls for {t2t_id}, "
            "produced by the T2T-ONT Snakemake workflow."
        ),
        "type":                "Genome Assembly",
        "research_data_files": ";".join(input_file_aliases),
    })

# --- ProcessDataFiles (hap1 + hap2 assembly + phased VCF per individual) ---
process_file_rows = []
for t2t_id in PUBLICATION_SAMPLES:
    for hap, fmt in [("hap1", "OTHER"), ("hap2", "OTHER"), ("variants", "VCF")]:  # FASTA not in GHGA CV
        fname = (
            f"{t2t_id}.assembly.{hap}.fasta" if hap in ("hap1", "hap2")
            else f"{t2t_id}.phased.vcf.gz"
        )
        process_file_rows.append({
            "alias":                  f"pfile_{t2t_id}_{hap}",
            "name":                   fname,
            "format":                 fmt,
            "analysis":               f"analysis_{t2t_id}",
            "dataset":                DATASET_ALIAS,
            "included_in_submission": True,
        })

# --- Dataset ----------------------------------------------------------------
dataset_rows = [{
    "alias":              DATASET_ALIAS,
    "title":              "T2T-ONT: ONT sequencing, assemblies and variant calls from 23 individuals",
    "description":        (
        "Ultra-long, Pore-C, and APK Oxford Nanopore raw sequencing data (unmapped BAMs), "
        "haplotype-resolved near-T2T assemblies (FASTA), and phased variant calls (VCF) "
        "for 23 healthy individuals. Generated as part of a study on single-platform "
        "ONT-based diploid T2T genome assembly."
    ),
    "types":              "Whole Genome Sequencing",
    "data_access_policy": DAP_ALIAS,
    "study":              STUDY_ALIAS,
}]

# --- DataAccessPolicy -------------------------------------------------------
dap_rows = [{
    "alias":                    DAP_ALIAS,
    "name":                     "T2T-ONT Data Access Policy",
    "description":              (
        "Policy governing access to human whole-genome sequencing data from the T2T-ONT study."
    ),
    "policy_text":              (
        "Your request will be evaluated by the responsible Data Access Committee (DAC). "
        "As soon as a positive decision is made, you will be provided with a data "
        "transfer agreement (DTA). After the DTA has been fully signed, access to the "
        "requested data is granted."
    ),
    "data_use_permission_term": "GENERAL_RESEARCH_USE",
    "data_use_permission_id":   "DUO:0000042",
    "data_use_modifier_terms":  "USER_SPECIFIC_RESTRICTION",
    "data_use_modifier_ids":    "DUO:0000026",
    "data_access_committee":    DAC_ALIAS,
}]

# --- DataAccessCommittee ----------------------------------------------------
dac_rows = [{
    "alias":     DAC_ALIAS,
    "email":     "imgag-dac@med.uni-tuebingen.de",
    "institute": "Institute of Medical Genetics and Applied Genomics, University of Tübingen",
}]

# --- Publication ------------------------------------------------------------
publication_rows = [{
    "alias":   "pub_T2T_ONT",
    "study":   STUDY_ALIAS,
    "title":   (
        "Single-Platform Nanopore Sequencing Enables Diploid Telomere-to-Telomere "
        "Genome Assembly and Haplotype-Resolved 3D Chromatin Maps"
    ),
    "author":  (
        "Gross C, Potabattula R, Cheng F, Leuchtenberg S, Hartung HS, "
        "Kristmann B, Buena-Atienza E, Casadei N, Ossowski S, Riess O"
    ),
    "year":    2026,
    "journal": "Nature Communications",
    "doi":     "",   # fill after publication
}]

# ---------------------------------------------------------------------------
# Write into a copy of the template (preserves all formatting)
# ---------------------------------------------------------------------------
import shutil

SHEET_DATA = {
    "Study":               study_rows,
    "Individual":          individual_rows,
    "Sample":              sample_rows,
    "ExperimentMethod":    method_rows,
    "Experiment":          experiment_rows,
    "ResearchDataFile":    file_rows,
    "AnalysisMethod":      analysis_method_rows,
    "Analysis":            analysis_rows,
    "ProcessDataFile":     process_file_rows,
    "Dataset":             dataset_rows,
    "DataAccessPolicy":    dap_rows,
    "DataAccessCommittee": dac_rows,
    "Publication":         publication_rows,
}

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)

wb = openpyxl.load_workbook(OUTPUT_FILE)

for sheet_name, data_rows in SHEET_DATA.items():
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: sheet '{sheet_name}' not in template, skipping")
        continue
    ws = wb[sheet_name]

    # Column-name → column-number from row 1
    col_index = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }

    # Clear existing data rows (keep header rows 1-6)
    for row in ws.iter_rows(min_row=HEADER_ROWS + 1):
        for cell in row:
            cell.value = None

    # Write data
    for r_idx, row_data in enumerate(data_rows):
        excel_row = HEADER_ROWS + 1 + r_idx
        for col_name, value in row_data.items():
            if col_name in col_index:
                ws.cell(row=excel_row, column=col_index[col_name], value=value)

    print(f"  {sheet_name:<22s}: {len(data_rows)} rows")

wb.save(OUTPUT_FILE)
print(f"\nSaved → {OUTPUT_FILE}")

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
by_method = {}
for e in experiment_rows:
    by_method[e["experiment_method"]] = by_method.get(e["experiment_method"], 0) + 1

missing = [f["name"] for f in file_rows if f["name"].startswith("MISSING_")]

print(f"\n--- Experiment counts ---")
for m, n in sorted(by_method.items()):
    print(f"  {m:<20s}: {n}")
print(f"  {'TOTAL':<20s}: {len(experiment_rows)}")

print(f"\n--- Pending fields ---")
print("  AnalysisMethod.workflow_version        (after tagging release)")
print("  Publication.doi                        (after publication)")
print("  ProcessDataFile.name (assembly/VCF)    (verify final filenames)")

if missing:
    print(f"\n--- WARNING: {len(missing)} BAM files not found on disk ---")
    for m in missing:
        print(f"  {m}")

# ---------------------------------------------------------------------------
# Generate upload file list TSV
# ---------------------------------------------------------------------------
FILE_LIST_OUT = BASE / "data/upload_folder/GHGA/ghga_upload_file_list.tsv"

upload_rows = []

# BAMs
for rid, (sample, bam_type) in rid_to_sample_type.items():
    bam_path = fc_to_path.get(rid)
    fname = fc_to_filename.get(rid, f"MISSING_{rid}.unmapped.bam")
    upload_rows.append({
        "sample":      sample,
        "category":    "Raw BAM",
        "type":        bam_type,
        "filename":    fname,
        "source_path": str(bam_path) if bam_path else "",
        "size_gb":     round(bam_path.stat().st_size / 1e9, 3) if bam_path else 0.0,
        "found":       bam_path is not None and bam_path.exists(),
    })

# Assembly FASTAs and phased VCFs
for sample in PUBLICATION_SAMPLES:
    for hap in ("hap1", "hap2"):
        src = BASE / f"assembly/output/final/{sample}/assembly.{hap}.pansn.fasta"
        upload_rows.append({
            "sample":      sample,
            "category":    "Assembly FASTA",
            "type":        hap,
            "filename":    src.name,
            "source_path": str(src),
            "size_gb":     round(src.stat().st_size / 1e9, 3) if src.exists() else 0.0,
            "found":       src.exists(),
        })
    src = BASE / f"results/{sample}/variants/small_variants.norm.vcf.gz"
    upload_rows.append({
        "sample":      sample,
        "category":    "Phased VCF",
        "type":        "-",
        "filename":    src.name,
        "source_path": str(src),
        "size_gb":     round(src.stat().st_size / 1e9, 3) if src.exists() else 0.0,
        "found":       src.exists(),
    })

fl_df = pd.DataFrame(upload_rows)
FILE_LIST_OUT.parent.mkdir(parents=True, exist_ok=True)
fl_df.to_csv(FILE_LIST_OUT, sep="\t", index=False)

total_gb = fl_df["size_gb"].sum()
not_found = fl_df[~fl_df["found"]]
print(f"\n--- Upload file list ---")
print(f"  Total files : {len(fl_df)}  ({total_gb:.1f} GB / {total_gb/1024:.2f} TB)")
for cat, grp in fl_df.groupby("category"):
    print(f"  {cat:<20}  {len(grp):>4} files   {grp['size_gb'].sum():>8.1f} GB")
if not not_found.empty:
    print(f"  WARNING: {len(not_found)} files not found on disk")
    for _, r in not_found.iterrows():
        print(f"    [{r['category']}] {r['sample']} / {r['filename']}")
print(f"  Saved → {FILE_LIST_OUT}")
